import requests
import openpyxl
import time
url = "https://api.opensea.io/api/v1/asset/0x2953399124f0cbb46d2cbacd8a89cf0599974963/1/"

headers = {"Accept": "application/json"}

response = requests.request("GET", url, headers=headers)

print(response.text)
def extract_token_details(search_term):

    parameters = {
        'fullText': 'true',
        'fields': 'capital;region;population'
    }

    response = requests.get(API_URL + search_term, params=parameters)

    json = response.json()

    if response.status_code != 200:
        return 'error'

    return json[0]


def write_token_details():
    filename = 'OpenseaTokens.xlsx'

    wb = openpyxl.load_workbook(filename)
    ws = wb.active
    wc = ws['A']

    i = 1
    for cell in wc[1:]:
        i += 1
        country_details = extract_country_details(cell.value)
        time.sleep(10)
        ws['B' + str(i)] = country_details['capital']
     

    wb.save(filename)